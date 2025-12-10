from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from reader.models import Meter, Bill


class ExportExcelView(APIView):
    """
    Exporta datos de facturas en un Excel con formato específico:
      - Si meter_type = 'WATER': Hoja "Agua" con facturas de agua
      - Si meter_type = 'ELECTRICITY': Hoja "Electricidad" con facturas de electricidad
      - Si meter_type = 'BOTH': Ambas hojas en un mismo Excel (requiere fechas)
      - Si meter_type = 'ALL': Histórico completo de ambos tipos (no requiere fechas)
      - Formato con celdas combinadas, estilos y colores
    """

    def get(self, request):
        meter_type = request.query_params.get('meter_type')  # 'WATER', 'ELECTRICITY', 'BOTH' o 'ALL'
        start_date = request.query_params.get('start_date')  # formato: YYYY-MM (opcional para ALL)
        end_date = request.query_params.get('end_date')      # formato: YYYY-MM (opcional para ALL)

        if not meter_type:
            return Response(
                {"detail": "Debe indicar 'meter_type'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar tipo de medidor (ahora incluye 'ALL' para histórico completo)
        valid_types = ['WATER', 'ELECTRICITY', 'BOTH', 'ALL']
        if meter_type not in valid_types:
            return Response(
                {"detail": "Tipo de medidor inválido. Use 'WATER', 'ELECTRICITY', 'BOTH' o 'ALL'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Si es ALL, no requiere fechas (exporta todo el histórico)
        if meter_type == 'ALL':
            start_period = 0  # Desde el inicio de los tiempos
            end_period = float('inf')  # Hasta el fin de los tiempos
            start_date = 'inicio'
            end_date = 'fin'
        else:
            # Validar que se proporcionen fechas para otros tipos
            if not start_date or not end_date:
                return Response(
                    {"detail": "Debe indicar 'start_date' y 'end_date' para este tipo de exportación."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Parseo de fechas: se trabaja con año y mes
            try:
                start_year, start_month = map(int, start_date.split('-'))
                end_year, end_month = map(int, end_date.split('-'))
            except ValueError:
                return Response(
                    {"detail": "Formato inválido de fechas. Use YYYY-MM."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            start_period = start_year * 12 + start_month
            end_period = end_year * 12 + end_month

        # Crear workbook
        output = BytesIO()
        workbook = Workbook()
        
        # Eliminar la hoja por defecto
        workbook.remove(workbook.active)

        # Crear hojas según el tipo de medidor seleccionado
        if meter_type == 'BOTH' or meter_type == 'ALL':
            # Exportar ambos tipos en hojas separadas
            # Hoja de Agua
            water_meters = Meter.objects.filter(meter_type='WATER')
            water_bills = Bill.objects.filter(meter__in=water_meters)
            water_bills = [b for b in water_bills if start_period <= (b.year * 12 + b.month) <= end_period]
            self._create_formatted_sheet(workbook, 'Agua', water_bills, 'Consumo [m3]')
            
            # Hoja de Electricidad
            electricity_meters = Meter.objects.filter(meter_type='ELECTRICITY')
            electricity_bills = Bill.objects.filter(meter__in=electricity_meters)
            electricity_bills = [b for b in electricity_bills if start_period <= (b.year * 12 + b.month) <= end_period]
            self._create_formatted_sheet(workbook, 'Electricidad', electricity_bills, 'Consumo [kWh]')
            
            if meter_type == 'ALL':
                filename = f"Facturas_Historico_Completo.xlsx"
            else:
                filename = f"Facturas_Completas_{start_date}_a_{end_date}.xlsx"
        else:
            # Exportar solo un tipo
            meters = Meter.objects.filter(meter_type=meter_type)
            bills = Bill.objects.filter(meter__in=meters)
            bills = [b for b in bills if start_period <= (b.year * 12 + b.month) <= end_period]

            if meter_type == 'WATER':
                self._create_formatted_sheet(workbook, 'Agua', bills, 'Consumo [m3]')
                filename = f"Facturas_AguasAndinas_{start_date}_a_{end_date}.xlsx"
            else:  # ELECTRICITY
                self._create_formatted_sheet(workbook, 'Electricidad', bills, 'Consumo [kWh]')
                filename = f"Facturas_Enel_{start_date}_a_{end_date}.xlsx"

        # Guardar y enviar el archivo
        workbook.save(output)
        output.seek(0)
            
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _get_unique_charges(self, bills):
        """
        Obtiene todos los cargos únicos de las boletas que tienen valor monetario o de consumo.
        Excluye solo cargos puramente informativos (textos, códigos, fechas sin monto asociado).
        Retorna una lista ordenada de nombres de cargos.
        """
        charge_names = set()
        
        for bill in bills:
            for charge in bill.charges.all():
                # Incluir solo cargos que:
                # 1. Tienen monto en $ (charge != 0), incluyendo negativos (descuentos), O
                # 2. Tienen valor numérico en m3/kWh (value > 0 y value_type es m3 o kWh)
                has_monetary_value = charge.charge != 0
                has_quantity_value = (
                    charge.value > 0 and 
                    charge.value_type in ['m3', 'kWh']
                )
                
                # Excluir SOLO cargos puramente informativos (sin ningún valor monetario/cantidad)
                # que son datos de contexto, no cargos reales
                is_informative = (
                    charge.value_type in ['código', 'fecha', 'texto', 'número'] and
                    charge.charge == 0
                ) or (
                    # Excluir tarifas unitarias (son informativas, no cargos aplicados)
                    charge.name.startswith('Tarifa') or
                    'Factor de cobro' in charge.name or
                    'Grupo tarifario' in charge.name or
                    'Último pago' in charge.name or
                    'Diámetro arranque' in charge.name
                )
                
                if (has_monetary_value or has_quantity_value) and not is_informative:
                    charge_names.add(charge.name)
        
        # Ordenar para tener consistencia
        return sorted(list(charge_names))

    def _create_formatted_sheet(self, workbook, sheet_name, bills, consumo_label):
        """
        Crea una hoja formateada con el estilo de la imagen de referencia.
        Incluye desagregación dinámica de cargos.
        """
        sheet = workbook.create_sheet(title=sheet_name)
        
        # Obtener todos los cargos únicos que tienen m3 o monto
        unique_charges = self._get_unique_charges(bills)
        num_charge_columns = len(unique_charges) * 2  # Cada cargo tiene 2 columnas (m3 y Monto)
        
        # Definir estilos
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Gris claro
        header_font = Font(name="Arial", bold=True, size=11)
        data_font = Font(name="Arial", size=11)
        center_alignment = Alignment(horizontal="center", vertical="center")
        bottom_center_alignment = Alignment(horizontal="center", vertical="bottom")
        
        # Bordes
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Borde derecho grueso para separar IDENTIFICACIÓN de CIFRAS DESTACADAS
        thick_right_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Borde inferior grueso para separar encabezados de datos
        thick_bottom_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thick', color='000000')
        )
        
        # Borde esquina (grueso derecho + grueso inferior)
        thick_corner_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thick', color='000000')
        )
        
        # Bordes para esquinas exteriores de la tabla
        top_left_corner = Border(
            left=Side(style='thick', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        top_right_corner = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Bordes superiores gruesos
        thick_top_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Borde izquierdo grueso
        thick_left_border = Border(
            left=Side(style='thick', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Borde derecho grueso (borde exterior derecho)
        thick_outer_right_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Calcular posiciones de columnas
        # Para AGUA: Columnas A-E (5 columnas), F-H (3 columnas), I+ (desagregación)
        # Para ELECTRICIDAD: Columnas A-F (6 columnas incluyendo Tarifa), G-I (3 columnas), J+ (desagregación)
        
        is_electricity = sheet_name == 'Electricidad'
        num_id_cols = 6 if is_electricity else 5
        
        last_cifras_col = num_id_cols + 3  # 8 para agua (E+3), 9 para electricidad (F+3)
        first_charge_col = last_cifras_col + 1
        last_charge_col = first_charge_col + num_charge_columns - 1
        
        # FILA 1-2: Encabezados principales combinados (abarcan filas 1 y 2)
        id_end_col = 'F' if is_electricity else 'E'
        cifras_start_col = 'G' if is_electricity else 'F'
        cifras_end_col = 'I' if is_electricity else 'H'
        
        sheet.merge_cells(f'A1:{id_end_col}2')  # IDENTIFICACIÓN
        sheet.merge_cells(f'{cifras_start_col}1:{cifras_end_col}2')  # CIFRAS DESTACADAS
        
        # Combinar celdas para DESAGREGACIÓN DE CARGOS si hay cargos (solo fila 1)
        if num_charge_columns > 0:
            from openpyxl.utils import get_column_letter
            start_col_letter = get_column_letter(first_charge_col)
            end_col_letter = get_column_letter(last_charge_col)
            sheet.merge_cells(f'{start_col_letter}1:{end_col_letter}1')
            sheet[f'{start_col_letter}1'] = 'Desagregación de Cargos'
            sheet[f'{start_col_letter}1'].fill = header_fill
            sheet[f'{start_col_letter}1'].font = header_font
            sheet[f'{start_col_letter}1'].alignment = bottom_center_alignment
        
        sheet['A1'] = 'IDENTIFICACIÓN'
        sheet[f'{cifras_start_col}1'] = 'CIFRAS DESTACADAS'
        
        # Altura de la fila 1 (doble altura)
        sheet.row_dimensions[1].height = 30
        
        # Aplicar estilos a encabezados principales con bordes exteriores gruesos
        sheet['A1'].fill = header_fill
        sheet['A1'].font = header_font
        sheet['A1'].alignment = bottom_center_alignment
        sheet['A1'].border = top_left_corner
        
        sheet[f'{cifras_start_col}1'].fill = header_fill
        sheet[f'{cifras_start_col}1'].font = header_font
        sheet[f'{cifras_start_col}1'].alignment = bottom_center_alignment
        sheet[f'{cifras_start_col}1'].border = thin_border
        
        # Aplicar borde superior grueso a las celdas intermedias de fila 1
        intermediate_cols = ['B', 'C', 'D'] if not is_electricity else ['B', 'C', 'D', 'E']
        for col in intermediate_cols:
            cell = sheet[f'{col}1']
            cell.border = thick_top_border
        
        # Última celda de IDENTIFICACIÓN tiene borde superior grueso y derecho grueso (división vertical)
        sheet[f'{id_end_col}1'].border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Celdas intermedias de CIFRAS DESTACADAS con borde superior grueso
        cifras_intermediate_cols = ['G', 'H'] if is_electricity else ['F', 'G']
        for col in cifras_intermediate_cols:
            cell = sheet[f'{col}1']
            cell.border = thick_top_border
        
        # Última celda de CIFRAS DESTACADAS - borde superior grueso y derecho grueso (división con desagregación)
        sheet[f'{cifras_end_col}1'].border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Aplicar bordes a las celdas de Desagregación de Cargos en fila 1
        if num_charge_columns > 0:
            from openpyxl.utils import get_column_letter
            for col_idx in range(first_charge_col, last_charge_col):
                col_letter = get_column_letter(col_idx)
                sheet[f'{col_letter}1'].border = thick_top_border
            
            # Última columna de desagregación (esquina superior derecha)
            last_col_letter = get_column_letter(last_charge_col)
            sheet[f'{last_col_letter}1'].border = top_right_corner
        else:
            # Si no hay cargos, última columna de CIFRAS es la esquina superior derecha
            sheet[f'{cifras_end_col}1'].border = top_right_corner
        
        # FILA 2: Sub-encabezados (nombres de cargos)
        # Para cada cargo, combinar 2 celdas para el nombre del cargo
        col_idx = first_charge_col
        for charge_name in unique_charges:
            from openpyxl.utils import get_column_letter
            start_col = get_column_letter(col_idx)
            end_col = get_column_letter(col_idx + 1)
            sheet.merge_cells(f'{start_col}2:{end_col}2')
            
            cell = sheet[f'{start_col}2']
            cell.value = charge_name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = bottom_center_alignment
            cell.border = thin_border
            
            col_idx += 2
        
        # FILA 3: Sub-sub-encabezados
        headers_row3 = [
            'ID Factura',
            'N° de Cliente',
        ]
        
        # Agregar Tarifa solo para electricidad
        if is_electricity:
            headers_row3.append('Tarifa')
        
        headers_row3.extend([
            'Macrozona',
            'Instalación',
            'Dirección',
            'Período',
            consumo_label,
            'Total a Pagar [$]'
        ])
        
        # Aplicar encabezados de IDENTIFICACIÓN y CIFRAS DESTACADAS en fila 3
        for col_num, header in enumerate(headers_row3, start=1):
            cell = sheet.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            
            # Calcular la última columna de IDENTIFICACIÓN y CIFRAS
            last_id_col_num = num_id_cols
            last_cifras_col_num = last_cifras_col
            
            # Aplicar bordes según la posición
            if col_num == 1:  # Primera columna - borde izquierdo grueso
                cell.border = Border(
                    left=Side(style='thick', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thick', color='000000')
                )
            elif col_num == last_id_col_num:  # Dirección (última columna de IDENTIFICACIÓN)
                cell.border = thick_corner_border  # Grueso abajo y derecho
            elif col_num == last_cifras_col_num:  # Total a Pagar (última columna de CIFRAS DESTACADAS)
                cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thick', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thick', color='000000')
                )
            else:
                cell.border = thick_bottom_border  # Solo grueso abajo
        
        # Aplicar sub-encabezados de unidad y "Monto [$]" para cada cargo en fila 3
        # Determinar unidad según tipo de hoja (m3 para agua, kWh/kW para electricidad)
        unit_header = 'm3' if sheet_name == 'Agua' else 'kWh/kW'
        
        col_idx = first_charge_col
        for charge_name in unique_charges:
            from openpyxl.utils import get_column_letter
            
            # Columna de unidad (m3 o kWh/kW)
            m3_col = get_column_letter(col_idx)
            cell_m3 = sheet[f'{m3_col}3']
            cell_m3.value = unit_header
            cell_m3.fill = header_fill
            cell_m3.font = header_font
            cell_m3.alignment = center_alignment
            cell_m3.border = thick_bottom_border
            
            # Columna Monto [$]
            monto_col = get_column_letter(col_idx + 1)
            cell_monto = sheet[f'{monto_col}3']
            cell_monto.value = 'Monto [$]'
            cell_monto.fill = header_fill
            cell_monto.font = header_font
            cell_monto.alignment = center_alignment
            
            # Si es la última columna, aplicar borde derecho grueso
            is_last_charge = (col_idx + 1 == last_charge_col)
            if is_last_charge:
                cell_monto.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thick', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thick', color='000000')
                )
            else:
                cell_monto.border = thick_bottom_border
            
            col_idx += 2
        
        # Ajustar borde derecho grueso en columnas de división
        # Última columna de IDENTIFICACIÓN (E para agua, F para electricidad)
        for row in [1, 2, 3]:
            sheet.cell(row=row, column=num_id_cols).border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thick', color='000000'),
                top=Side(style='thick' if row == 1 else 'thin', color='000000'),
                bottom=Side(style='thick' if row == 3 else 'thin', color='000000')
            )
        
        # Última columna de CIFRAS DESTACADAS (H para agua, I para electricidad)
        for row in [1, 2, 3]:
            sheet.cell(row=row, column=last_cifras_col).border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thick', color='000000'),
                top=Side(style='thick' if row == 1 else 'thin', color='000000'),
                bottom=Side(style='thick' if row == 3 else 'thin', color='000000')
            )
        
        # Ajustar ancho de columnas
        sheet.column_dimensions['A'].width = 12  # ID Factura
        sheet.column_dimensions['B'].width = 16  # N° de Cliente
        
        if is_electricity:
            sheet.column_dimensions['C'].width = 20  # Tarifa
            sheet.column_dimensions['D'].width = 14  # Macrozona
            sheet.column_dimensions['E'].width = 18  # Instalación
            sheet.column_dimensions['F'].width = 35  # Dirección
            sheet.column_dimensions['G'].width = 12  # Período
            sheet.column_dimensions['H'].width = 16  # Consumo
            sheet.column_dimensions['I'].width = 20  # Total a Pagar
        else:
            sheet.column_dimensions['C'].width = 14  # Macrozona
            sheet.column_dimensions['D'].width = 18  # Instalación
            sheet.column_dimensions['E'].width = 35  # Dirección
            sheet.column_dimensions['F'].width = 12  # Período
            sheet.column_dimensions['G'].width = 16  # Consumo
            sheet.column_dimensions['H'].width = 20  # Total a Pagar
        
        # Ajustar ancho de columnas de desagregación
        from openpyxl.utils import get_column_letter
        for col_idx in range(first_charge_col, last_charge_col + 1):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = 15  # Aumentado de 12 a 15
        
        # FILA 4+: Datos de las facturas
        row_num = 4
        total_rows = len(bills)
        
        for idx, bill in enumerate(bills):
            # Determinar si es la última fila
            is_last_row = (idx == total_rows - 1)
            
            # Formatear período como "MM/YYYY"
            periodo = f"{bill.month:02d}/{bill.year}"
            
            # Obtener el valor de consumo desde los cargos
            consumo_value = ''
            if sheet_name == 'Electricidad':
                cargo_consumo = bill.charges.filter(name__icontains='Electricidad Consumida').first()
                if cargo_consumo:
                    consumo_value = float(cargo_consumo.value)
            else:  # Agua
                cargo_consumo = bill.charges.filter(name__icontains='CONSUMO AGUA').first()
                if cargo_consumo:
                    consumo_value = float(cargo_consumo.value)
            
            # Datos de IDENTIFICACIÓN y CIFRAS DESTACADAS
            data_row = [
                bill.invoice_number,  # ID Factura
                bill.meter.client_number,
            ]
            
            # Agregar Tarifa solo para electricidad
            if is_electricity:
                data_row.append(bill.tarifa)
            
            data_row.extend([
                bill.meter.macrozona,
                bill.meter.instalacion,
                bill.meter.direccion,
                periodo,
                consumo_value,
                int(bill.total_to_pay)  # Convertir a entero (sin decimales)
            ])

            # Escribir datos de IDENTIFICACIÓN y CIFRAS DESTACADAS
            for col_num, value in enumerate(data_row, start=1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.font = data_font
                
                # Aplicar bordes según la posición
                if col_num == 1:  # Primera columna
                    if is_last_row:
                        cell.border = Border(
                            left=Side(style='thick', color='000000'),
                            right=Side(style='thin', color='000000'),
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thick', color='000000')
                        )
                    else:
                        cell.border = thick_left_border
                elif col_num == num_id_cols:  # Dirección (última col de IDENTIFICACIÓN) - división vertical
                    if is_last_row:
                        cell.border = Border(
                            left=Side(style='thin', color='000000'),
                            right=Side(style='thick', color='000000'),
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thick', color='000000')
                        )
                    else:
                        cell.border = thick_right_border
                elif col_num == last_cifras_col:  # Total a Pagar - división vertical
                    if is_last_row:
                        cell.border = Border(
                            left=Side(style='thin', color='000000'),
                            right=Side(style='thick', color='000000'),
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thick', color='000000')
                        )
                    else:
                        cell.border = thick_right_border
                else:
                    if is_last_row:
                        cell.border = Border(
                            left=Side(style='thin', color='000000'),
                            right=Side(style='thin', color='000000'),
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thick', color='000000')
                        )
                    else:
                        cell.border = thin_border
                
                # Alineación
                periodo_col = num_id_cols + 1
                consumo_col = num_id_cols + 2
                total_col = num_id_cols + 3
                
                if col_num in [1, periodo_col, consumo_col, total_col]:  # ID, Período, Consumo, Total - centrado
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Formato de número
                if col_num == consumo_col and consumo_value:  # Consumo
                    cell.number_format = '#,##0.00'
                elif col_num == total_col:  # Total a Pagar (dinero - sin decimales)
                    cell.number_format = '#,##0'
            
            # Escribir datos de DESAGREGACIÓN DE CARGOS
            col_idx = first_charge_col
            for charge_name in unique_charges:
                from openpyxl.utils import get_column_letter
                
                # Buscar el cargo correspondiente en esta boleta
                charge = bill.charges.filter(name=charge_name).first()
                
                # Valor de m3/kWh
                m3_value = ''
                if charge and charge.value > 0 and charge.value_type in ['m3', 'kWh']:
                    m3_value = float(charge.value)
                
                # Valor de Monto [$] (incluye negativos como descuentos)
                monto_value = ''
                if charge and charge.charge != 0:
                    monto_value = int(charge.charge)  # Convertir a entero (sin decimales)
                
                # Escribir m3
                m3_col = get_column_letter(col_idx)
                cell_m3 = sheet[f'{m3_col}{row_num}']
                cell_m3.value = m3_value
                cell_m3.font = data_font
                cell_m3.alignment = Alignment(horizontal="center", vertical="center")
                if m3_value:
                    cell_m3.number_format = '#,##0.00'
                
                if is_last_row:
                    cell_m3.border = Border(
                        left=Side(style='thin', color='000000'),
                        right=Side(style='thin', color='000000'),
                        top=Side(style='thin', color='000000'),
                        bottom=Side(style='thick', color='000000')
                    )
                else:
                    cell_m3.border = thin_border
                
                # Escribir Monto [$]
                monto_col = get_column_letter(col_idx + 1)
                cell_monto = sheet[f'{monto_col}{row_num}']
                cell_monto.value = monto_value
                cell_monto.font = data_font
                cell_monto.alignment = Alignment(horizontal="center", vertical="center")
                if monto_value:
                    cell_monto.number_format = '#,##0'  # Sin decimales para dinero
                
                # Si es la última columna, aplicar borde derecho grueso
                is_last_charge = (col_idx + 1 == last_charge_col)
                if is_last_charge:
                    if is_last_row:
                        cell_monto.border = Border(
                            left=Side(style='thin', color='000000'),
                            right=Side(style='thick', color='000000'),
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thick', color='000000')
                        )
                    else:
                        cell_monto.border = thick_outer_right_border
                else:
                    if is_last_row:
                        cell_monto.border = Border(
                            left=Side(style='thin', color='000000'),
                            right=Side(style='thin', color='000000'),
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thick', color='000000')
                        )
                    else:
                        cell_monto.border = thin_border
                
                col_idx += 2
            
            row_num += 1
