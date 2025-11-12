from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from reader.models import Meter, Bill


class ExportExcelView(APIView):
    """
    Exporta datos de facturas en un Excel con formato específico:
      - Si meter_type = 'WATER': Hoja "Agua" con facturas de agua
      - Si meter_type = 'ELECTRICITY': Hoja "Electricidad" con facturas de electricidad
      - Si meter_type = 'BOTH': Ambas hojas en un mismo Excel (requiere fechas)
      - Si meter_type = 'ALL': Histórico completo de ambos tipos (no requiere fechas)
      - Formato con celdas combinadas, estilos y colores
    """

    def get(self, request):
        meter_type = request.query_params.get('meter_type')  # 'WATER', 'ELECTRICITY', 'BOTH' o 'ALL'
        start_date = request.query_params.get('start_date')  # formato: YYYY-MM (opcional para ALL)
        end_date = request.query_params.get('end_date')      # formato: YYYY-MM (opcional para ALL)

        if not meter_type:
            return Response(
                {"detail": "Debe indicar 'meter_type'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar tipo de medidor (ahora incluye 'ALL' para histórico completo)
        valid_types = ['WATER', 'ELECTRICITY', 'BOTH', 'ALL']
        if meter_type not in valid_types:
            return Response(
                {"detail": "Tipo de medidor inválido. Use 'WATER', 'ELECTRICITY', 'BOTH' o 'ALL'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Si es ALL, no requiere fechas (exporta todo el histórico)
        if meter_type == 'ALL':
            start_period = 0  # Desde el inicio de los tiempos
            end_period = float('inf')  # Hasta el fin de los tiempos
            start_date = 'inicio'
            end_date = 'fin'
        else:
            # Validar que se proporcionen fechas para otros tipos
            if not start_date or not end_date:
                return Response(
                    {"detail": "Debe indicar 'start_date' y 'end_date' para este tipo de exportación."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Parseo de fechas: se trabaja con año y mes
            try:
                start_year, start_month = map(int, start_date.split('-'))
                end_year, end_month = map(int, end_date.split('-'))
            except ValueError:
                return Response(
                    {"detail": "Formato inválido de fechas. Use YYYY-MM."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            start_period = start_year * 12 + start_month
            end_period = end_year * 12 + end_month

        # Crear workbook
        output = BytesIO()
        workbook = Workbook()
        
        # Eliminar la hoja por defecto
        workbook.remove(workbook.active)

        # Crear hojas según el tipo de medidor seleccionado
        if meter_type == 'BOTH' or meter_type == 'ALL':
            # Exportar ambos tipos en hojas separadas
            # Hoja de Agua
            water_meters = Meter.objects.filter(meter_type='WATER')
            water_bills = Bill.objects.filter(meter__in=water_meters)
            water_bills = [b for b in water_bills if start_period <= (b.year * 12 + b.month) <= end_period]
            self._create_formatted_sheet(workbook, 'Agua', water_bills, 'Consumo [m3]')
            
            # Hoja de Electricidad
            electricity_meters = Meter.objects.filter(meter_type='ELECTRICITY')
            electricity_bills = Bill.objects.filter(meter__in=electricity_meters)
            electricity_bills = [b for b in electricity_bills if start_period <= (b.year * 12 + b.month) <= end_period]
            self._create_formatted_sheet(workbook, 'Electricidad', electricity_bills, 'Consumo [kWh]')
            
            if meter_type == 'ALL':
                filename = f"Facturas_Historico_Completo.xlsx"
            else:
                filename = f"Facturas_Completas_{start_date}_a_{end_date}.xlsx"
        else:
            # Exportar solo un tipo
            meters = Meter.objects.filter(meter_type=meter_type)
            bills = Bill.objects.filter(meter__in=meters)
            bills = [b for b in bills if start_period <= (b.year * 12 + b.month) <= end_period]

            if meter_type == 'WATER':
                self._create_formatted_sheet(workbook, 'Agua', bills, 'Consumo [m3]')
                filename = f"Facturas_AguasAndinas_{start_date}_a_{end_date}.xlsx"
            else:  # ELECTRICITY
                self._create_formatted_sheet(workbook, 'Electricidad', bills, 'Consumo [kWh]')
                filename = f"Facturas_Enel_{start_date}_a_{end_date}.xlsx"

        # Guardar y enviar el archivo
        workbook.save(output)
        output.seek(0)
            
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    def _create_formatted_sheet(self, workbook, sheet_name, bills, consumo_label):
        """
        Crea una hoja formateada con el estilo de la imagen de referencia.
        """
        sheet = workbook.create_sheet(title=sheet_name)
        
        # Definir estilos
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Gris claro
        header_font = Font(name="Arial", bold=True, size=11)
        data_font = Font(name="Arial", size=11)
        center_alignment = Alignment(horizontal="center", vertical="center")
        bottom_alignment = Alignment(horizontal="center", vertical="bottom")
        
        # Bordes
        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Borde derecho grueso para separar IDENTIFICACIÓN de CIFRAS DESTACADAS
        thick_right_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Borde inferior grueso para separar encabezados de datos
        thick_bottom_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thick', color='000000')
        )
        
        # Borde esquina (grueso derecho + grueso inferior)
        thick_corner_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thick', color='000000')
        )
        
        # Bordes para esquinas exteriores de la tabla
        top_left_corner = Border(
            left=Side(style='thick', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        top_right_corner = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Bordes superiores gruesos
        thick_top_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Borde izquierdo grueso
        thick_left_border = Border(
            left=Side(style='thick', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Borde derecho grueso (borde exterior derecho)
        thick_outer_right_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # FILA 1: Encabezados principales combinados
        sheet.merge_cells('A1:E1')  # IDENTIFICACIÓN
        sheet.merge_cells('F1:H1')  # CIFRAS DESTACADAS
        
        sheet['A1'] = 'IDENTIFICACIÓN'
        sheet['F1'] = 'CIFRAS DESTACADAS'
        
        # Altura de la fila 1 (doble altura)
        sheet.row_dimensions[1].height = 30
        
        # Aplicar estilos a encabezados principales con bordes exteriores gruesos
        sheet['A1'].fill = header_fill
        sheet['A1'].font = header_font
        sheet['A1'].alignment = bottom_alignment  # Alineado abajo
        sheet['A1'].border = top_left_corner  # Esquina superior izquierda
        
        sheet['F1'].fill = header_fill
        sheet['F1'].font = header_font
        sheet['F1'].alignment = bottom_alignment  # Alineado abajo
        sheet['F1'].border = thin_border
        
        # Aplicar borde superior grueso a las celdas intermedias de fila 1
        for col in ['B', 'C', 'D']:
            cell = sheet[f'{col}1']
            cell.border = thick_top_border
        
        # Celda E1 tiene borde superior grueso y derecho grueso (división vertical)
        sheet['E1'].border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thick', color='000000'),
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        # Celdas F1 y G1 con borde superior grueso
        for col in ['F', 'G']:
            cell = sheet[f'{col}1']
            cell.border = thick_top_border
        
        # Celda H1 - esquina superior derecha
        sheet['H1'].border = top_right_corner
        
        # FILA 2: Sub-encabezados
        headers = [
            'ID Factura',
            'N° de Cliente',
            'Macrozona',
            'Instalación',
            'Dirección',
            'Período',
            consumo_label,  # 'Consumo [m3]' o 'Consumo [kWh]'
            'Total a Pagar [$]'
        ]
        
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            
            # Aplicar bordes según la posición
            if col_num == 1:  # Primera columna - borde izquierdo grueso
                cell.border = Border(
                    left=Side(style='thick', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thick', color='000000')
                )
            elif col_num == 5:  # Dirección (última columna de IDENTIFICACIÓN)
                cell.border = thick_corner_border  # Grueso abajo y derecho
            elif col_num == 8:  # Última columna - borde derecho grueso
                cell.border = Border(
                    left=Side(style='thin', color='000000'),
                    right=Side(style='thick', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thick', color='000000')
                )
            else:
                cell.border = thick_bottom_border  # Solo grueso abajo
        
        # Aplicar borde derecho grueso a las celdas de la columna E en fila 1
        # (ya combinadas, pero necesitamos asegurar el borde vertical)
        for row in [1, 2]:
            sheet.cell(row=row, column=5).border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thick', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thick' if row == 2 else 'thin', color='000000')
            )
        
        # Ajustar ancho de columnas
        sheet.column_dimensions['A'].width = 12  # ID Factura
        sheet.column_dimensions['B'].width = 15  # N° de Cliente
        sheet.column_dimensions['C'].width = 12  # Macrozona
        sheet.column_dimensions['D'].width = 15  # Instalación
        sheet.column_dimensions['E'].width = 20  # Dirección
        sheet.column_dimensions['F'].width = 12  # Período
        sheet.column_dimensions['G'].width = 15  # Consumo
        sheet.column_dimensions['H'].width = 18  # Total a Pagar
        
        # FILA 3+: Datos de las facturas
        row_num = 3
        total_rows = len(bills)
        
        for idx, bill in enumerate(bills):
            # Formatear período como "MM/YYYY"
            periodo = f"{bill.month:02d}/{bill.year}"
            
            # Obtener el valor de consumo desde los cargos
            consumo_value = ''
            if sheet_name == 'Electricidad':
                # Buscar el cargo "Electricidad Consumida"
                cargo_consumo = bill.charges.filter(name__icontains='Electricidad Consumida').first()
                if cargo_consumo:
                    consumo_value = float(cargo_consumo.value)
            else:  # Agua
                # Buscar el cargo "CONSUMO AGUA"
                cargo_consumo = bill.charges.filter(name__icontains='CONSUMO AGUA').first()
                if cargo_consumo:
                    consumo_value = float(cargo_consumo.value)
            
            data_row = [
                '',  # ID Factura - dejar en blanco
                bill.meter.client_number,
                bill.meter.macrozona,
                bill.meter.instalacion,
                bill.meter.direccion,
                periodo,
                consumo_value,  # Consumo obtenido de cargos
                float(bill.total_to_pay)
            ]
            
            is_last_row = (idx == total_rows - 1)
            
            for col_num, value in enumerate(data_row, start=1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.font = data_font
                
                # Aplicar bordes según la posición
                if col_num == 1:  # Primera columna - borde izquierdo grueso
                    if is_last_row:
                        cell.border = Border(
                            left=Side(style='thick', color='000000'),
                            right=Side(style='thin', color='000000'),
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thick', color='000000')
                        )
                    else:
                        cell.border = thick_left_border
                elif col_num == 5:  # Dirección - borde derecho grueso (división vertical)
                    if is_last_row:
                        cell.border = Border(
                            left=Side(style='thin', color='000000'),
                            right=Side(style='thick', color='000000'),
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thick', color='000000')
                        )
                    else:
                        cell.border = thick_right_border
                elif col_num == 8:  # Última columna - borde derecho grueso
                    if is_last_row:
                        cell.border = Border(
                            left=Side(style='thin', color='000000'),
                            right=Side(style='thick', color='000000'),
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thick', color='000000')
                        )
                    else:
                        cell.border = thick_outer_right_border
                else:
                    if is_last_row:  # Última fila - borde inferior grueso
                        cell.border = Border(
                            left=Side(style='thin', color='000000'),
                            right=Side(style='thin', color='000000'),
                            top=Side(style='thin', color='000000'),
                            bottom=Side(style='thick', color='000000')
                        )
                    else:
                        cell.border = thin_border
                
                # Alineación
                if col_num in [1, 6, 7, 8]:  # ID, Período, Consumo, Total - centrado
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                
                # Formato de número
                if col_num == 7 and consumo_value:  # Consumo
                    cell.number_format = '#,##0.00'
                elif col_num == 8:  # Total a Pagar
                    cell.number_format = '#,##0.00'
            
            row_num += 1
